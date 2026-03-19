from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
import openpyxl
import csv
import io
from app.core.deps import get_current_user
from app.core.database import db
from app.models.user import UserInDB
from app.models.product import ProductCreate, ProductUpdate, ProductResponse, ProductListResponse, ProductInDB
from bson import ObjectId

router = APIRouter()


# --- Products Excel/CSV Import Endpoint ---
@router.post("/import-excel")
async def import_products_excel(
    file: UploadFile = File(...),
    current_user: UserInDB = Depends(get_current_user)
):
    # File validation
    allowed_ext = [".xlsx", ".xls", ".csv"]
    filename = file.filename.lower()
    if not any(filename.endswith(ext) for ext in allowed_ext):
        raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx, .xls, .csv allowed.")
    if file.size and file.size > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5MB)")

    # Get shopId
    shop = await db.get_db().shops.find_one({"userId": str(current_user.id)})
    if not shop:
        raise HTTPException(status_code=400, detail="Shop profile must be created first")
    shop_id = str(shop["_id"])

    # Read file content
    content = await file.read()
    imported, updated, errors = 0, 0, []
    total_rows = 0
    try:
        if filename.endswith(".csv"):
            decoded = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        else:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    # Column mapping (case-insensitive)
    col_map = {
        "name": ["name", "product name", "item"],
        "price": ["price", "cost", "rate"],
        "description": ["description", "details", "about"],
        "stock": ["stock", "quantity", "qty"],
        "category": ["category", "type"]
    }
    def get_col(row, keys, default=None):
        for k in keys:
            for col in row:
                if col and str(col).strip().lower() == k:
                    return row[col]
        return default

    for idx, row in enumerate(rows, start=2):
        total_rows += 1
        try:
            # Flexible column matching
            name = get_col(row, col_map["name"])
            price = get_col(row, col_map["price"])
            description = get_col(row, col_map["description"], "")
            stock = get_col(row, col_map["stock"], 0)
            category = get_col(row, col_map["category"], "General")
            if not name:
                raise ValueError("Missing product name")
            try:
                price = float(price) if price is not None else 0.0
            except:
                price = 0.0
            try:
                stock = int(stock) if stock is not None else 0
            except:
                stock = 0
            # Upsert by name+shopId
            existing = await db.get_db().products.find_one({"shopId": shop_id, "name": name})
            product_data = {
                "shopId": shop_id,
                "name": name,
                "price": price,
                "description": description or "",
                "stock": stock,
                "category": category or "General",
                "updatedAt": datetime.utcnow(),
            }
            if existing:
                await db.get_db().products.update_one({"_id": existing["_id"]}, {"$set": product_data})
                updated += 1
            else:
                product_data["createdAt"] = datetime.utcnow()
                await db.get_db().products.insert_one(product_data)
                imported += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")

    return {
        "imported": imported,
        "updated": updated,
        "errors": errors,
        "total_rows": total_rows
    }

@router.get("/", response_model=ProductListResponse)
async def list_products(
    page: int = 1,
    limit: int = 8,
    category: Optional[str] = None,
    sort: Optional[str] = None, # NEWEST_FIRST, PRICE_HIGH, PRICE_LOW
    current_user: UserInDB = Depends(get_current_user)
):
    skip = (page - 1) * limit
    
    # Get shop ID for current user
    shop = await db.get_db().shops.find_one({"userId": str(current_user.id)})
    if not shop:
        return {"products": [], "total": 0, "page": page, "limit": limit}
    
    shop_id = str(shop["_id"])
    query = {"shopId": shop_id}
    if category:
        query["category"] = category
        
    cursor = db.get_db().products.find(query)
    
    # Sorting
    if sort == "PRICE_HIGH":
        cursor.sort("price", -1)
    elif sort == "PRICE_LOW":
        cursor.sort("price", 1)
    else: # NEWEST_FIRST or default
        cursor.sort("createdAt", -1)
        
    total = await db.get_db().products.count_documents(query)
    cursor.skip(skip).limit(limit)
    
    products = []
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        doc["shopId"] = str(doc["shopId"])
        products.append(ProductResponse(**doc))
        
    return {
        "products": products,
        "total": total,
        "page": page,
        "limit": limit
    }

@router.post("/", response_model=ProductResponse)
async def create_product(
    product_in: ProductCreate,
    current_user: UserInDB = Depends(get_current_user)
):
    shop = await db.get_db().shops.find_one({"userId": str(current_user.id)})
    if not shop:
        # Auto-create shop if missing? Or error. Let's error for strictness.
        raise HTTPException(status_code=400, detail="Shop profile must be created first")
    
    shop_id = str(shop["_id"])
    
    # Create product document for database
    product_data = product_in.model_dump(by_alias=True)
    product_data["shopId"] = shop_id
    product_data["createdAt"] = datetime.utcnow()
    product_data["updatedAt"] = datetime.utcnow()
    
    result = await db.get_db().products.insert_one(product_data)
    
    # Fetch the created product and return it
    created_product = await db.get_db().products.find_one({"_id": result.inserted_id})
    created_product["_id"] = str(created_product["_id"])
    created_product["shopId"] = str(created_product["shopId"])
    
    return ProductResponse(**created_product)

@router.put("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str,
    product_in: ProductUpdate,
    current_user: UserInDB = Depends(get_current_user)
):
    # Verify ownership
    try:
        product = await db.get_db().products.find_one({"_id": ObjectId(product_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid product ID")
        
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    shop = await db.get_db().shops.find_one({"userId": str(current_user.id)})
    if not shop or str(product["shopId"]) != str(shop["_id"]):
        raise HTTPException(status_code=403, detail="Not authorized to update this product")
        
    update_data = product_in.model_dump(exclude_unset=True, by_alias=True)
    update_data["updatedAt"] = datetime.utcnow()
    
    await db.get_db().products.update_one(
        {"_id": ObjectId(product_id)},
        {"$set": update_data}
    )
    
    updated_product = await db.get_db().products.find_one({"_id": ObjectId(product_id)})
    updated_product["_id"] = str(updated_product["_id"])
    updated_product["shopId"] = str(updated_product["shopId"])
    return ProductResponse(**updated_product)

@router.delete("/{product_id}")
async def delete_product(
    product_id: str,
    current_user: UserInDB = Depends(get_current_user)
):
    # Verify ownership
    try:
        product = await db.get_db().products.find_one({"_id": ObjectId(product_id)})
    except:
        raise HTTPException(status_code=400, detail="Invalid product ID")
        
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    shop = await db.get_db().shops.find_one({"userId": str(current_user.id)})
    if not shop or str(product["shopId"]) != str(shop["_id"]):
        raise HTTPException(status_code=403, detail="Not authorized to delete this product")

    await db.get_db().products.delete_one({"_id": ObjectId(product_id)})
    return {"message": "Product deleted successfully"}
